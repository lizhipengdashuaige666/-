"""Payment business logic — CRUD, status transitions, Excel import/export.

Pure Python, no Qt dependencies. All database operations go through this class.
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable

import openpyxl

from unified.payment_db import DB_DIR, get_db
from unified.payment_models import (
    ImportResult,
    PaymentAttachment,
    PaymentLog,
    PaymentRecord,
    PaymentStatus,
    STATUS_DISPLAY,
    TRANSITIONS,
)


class PaymentService:
    def __init__(self) -> None:
        self._db = get_db()

    # ── CRUD ──────────────────────────────────────────────────────────

    def get_all(
        self, limit: int = 200, offset: int = 0,
        sort_by: str = "updated_time", sort_desc: bool = True,
    ) -> list[PaymentRecord]:
        allowed = {"created_time", "updated_time", "apply_date", "amount", "supplier_name"}
        col = sort_by if sort_by in allowed else "updated_time"
        direction = "DESC" if sort_desc else "ASC"
        cur = self._db.execute(
            f"SELECT id, payment_no, supplier_name, amount, apply_date, po_number, "
            f"status, notes, created_time, updated_time "
            f"FROM payment_records ORDER BY {col} {direction} LIMIT ? OFFSET ?",
            (limit, offset),
        )
        return [PaymentRecord(*row) for row in cur.fetchall()]

    def get_by_id(self, payment_id: int) -> PaymentRecord | None:
        cur = self._db.execute(
            "SELECT id, payment_no, supplier_name, amount, apply_date, po_number, "
            "status, notes, created_time, updated_time "
            "FROM payment_records WHERE id = ?",
            (payment_id,),
        )
        row = cur.fetchone()
        return PaymentRecord(*row) if row else None

    def get_by_payment_no(self, payment_no: str) -> PaymentRecord | None:
        cur = self._db.execute(
            "SELECT id, payment_no, supplier_name, amount, apply_date, po_number, "
            "status, notes, created_time, updated_time "
            "FROM payment_records WHERE payment_no = ?",
            (payment_no,),
        )
        row = cur.fetchone()
        return PaymentRecord(*row) if row else None

    def count_by_status(self) -> dict[str, int]:
        cur = self._db.execute(
            "SELECT status, COUNT(*) FROM payment_records GROUP BY status"
        )
        result: dict[str, int] = {s.value: 0 for s in PaymentStatus}
        for status, count in cur.fetchall():
            result[status] = count
        return result

    def total_count(self) -> int:
        cur = self._db.execute("SELECT COUNT(*) FROM payment_records")
        return cur.fetchone()[0]

    # ── Status transitions ────────────────────────────────────────────

    def transition_status(
        self, payment_id: int, new_status: str,
        operator: str = "当前用户", notes: str = "",
    ) -> PaymentRecord:
        record = self.get_by_id(payment_id)
        if record is None:
            raise ValueError(f"付款记录不存在: id={payment_id}")

        current = PaymentStatus(record.status)
        target = PaymentStatus(new_status)
        allowed = TRANSITIONS.get(current, [])
        if target not in allowed:
            raise ValueError(
                f"不允许从「{current.value}」直接变更为「{target.value}」"
            )

        self._db.execute(
            "UPDATE payment_records SET status=?, updated_time=datetime('now','localtime') WHERE id=?",
            (new_status, payment_id),
        )
        action = f"状态变更: {current.value} → {target.value}"
        self._db.execute(
            "INSERT INTO payment_logs (payment_id, action, old_status, new_status, operator, notes) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (payment_id, action, record.status, new_status, operator, notes),
        )
        self._db.commit()
        return self.get_by_id(payment_id)

    # ── Logs ──────────────────────────────────────────────────────────

    def get_logs_for_payment(self, payment_id: int) -> list[PaymentLog]:
        cur = self._db.execute(
            "SELECT id, payment_id, action, old_status, new_status, operator, timestamp, notes "
            "FROM payment_logs WHERE payment_id = ? ORDER BY timestamp ASC",
            (payment_id,),
        )
        return [PaymentLog(*row) for row in cur.fetchall()]

    def get_recent_activity(self, limit: int = 20) -> list[PaymentLog]:
        cur = self._db.execute(
            "SELECT id, payment_id, action, old_status, new_status, operator, timestamp, notes "
            "FROM payment_logs ORDER BY timestamp DESC LIMIT ?",
            (limit,),
        )
        return [PaymentLog(*row) for row in cur.fetchall()]

    # ── Search & Filter ───────────────────────────────────────────────

    def search(
        self, query: str = "", filters: dict | None = None,
        limit: int = 200, offset: int = 0,
        sort_by: str = "updated_time", sort_desc: bool = True,
    ) -> list[PaymentRecord]:
        allowed_cols = {"created_time", "updated_time", "apply_date", "amount", "supplier_name"}
        col = sort_by if sort_by in allowed_cols else "updated_time"
        direction = "DESC" if sort_desc else "ASC"

        conditions: list[str] = []
        params: list = []

        if query:
            like = f"%{query}%"
            conditions.append(
                "(payment_no LIKE ? OR supplier_name LIKE ? OR po_number LIKE ?)"
            )
            params.extend([like, like, like])

        if filters:
            if "status" in filters and filters["status"]:
                statuses = filters["status"]
                if isinstance(statuses, str):
                    statuses = [statuses]
                placeholders = ",".join("?" for _ in statuses)
                conditions.append(f"status IN ({placeholders})")
                params.extend(statuses)

            if "date_from" in filters and filters["date_from"]:
                conditions.append("apply_date >= ?")
                params.append(filters["date_from"])

            if "date_to" in filters and filters["date_to"]:
                conditions.append("apply_date <= ?")
                params.append(filters["date_to"])

            if "amount_min" in filters and filters["amount_min"] is not None:
                conditions.append("amount >= ?")
                params.append(filters["amount_min"])

            if "amount_max" in filters and filters["amount_max"] is not None:
                conditions.append("amount <= ?")
                params.append(filters["amount_max"])

        where = (" AND ".join(conditions)) if conditions else "1=1"
        sql = (
            f"SELECT id, payment_no, supplier_name, amount, apply_date, po_number, "
            f"status, notes, created_time, updated_time "
            f"FROM payment_records WHERE {where} "
            f"ORDER BY {col} {direction} LIMIT ? OFFSET ?"
        )
        params.extend([limit, offset])
        cur = self._db.execute(sql, params)
        return [PaymentRecord(*row) for row in cur.fetchall()]

    def search_count(self, query: str = "", filters: dict | None = None) -> int:
        conditions: list[str] = []
        params: list = []

        if query:
            like = f"%{query}%"
            conditions.append(
                "(payment_no LIKE ? OR supplier_name LIKE ? OR po_number LIKE ?)"
            )
            params.extend([like, like, like])

        if filters:
            if "status" in filters and filters["status"]:
                statuses = filters["status"]
                if isinstance(statuses, str):
                    statuses = [statuses]
                placeholders = ",".join("?" for _ in statuses)
                conditions.append(f"status IN ({placeholders})")
                params.extend(statuses)

            if "date_from" in filters and filters["date_from"]:
                conditions.append("apply_date >= ?")
                params.append(filters["date_from"])

            if "date_to" in filters and filters["date_to"]:
                conditions.append("apply_date <= ?")
                params.append(filters["date_to"])

        where = (" AND ".join(conditions)) if conditions else "1=1"
        cur = self._db.execute(
            f"SELECT COUNT(*) FROM payment_records WHERE {where}", params
        )
        return cur.fetchone()[0]

    # ── Dashboard stats ───────────────────────────────────────────────

    def get_dashboard_stats(self) -> dict:
        by_status = self.count_by_status()
        now = datetime.now()
        today_str = now.strftime("%Y-%m-%d")
        month_start = now.strftime("%Y-%m-01")

        cur = self._db.execute(
            "SELECT COUNT(*) FROM payment_records WHERE status='已付款' AND updated_time >= ?",
            (today_str,),
        )
        paid_today = cur.fetchone()[0]

        cur = self._db.execute(
            "SELECT COUNT(*) FROM payment_records WHERE status='已付款' AND updated_time >= ?",
            (month_start,),
        )
        paid_this_month = cur.fetchone()[0]

        overdue = self.scan_overdue()

        cur = self._db.execute(
            "SELECT id, payment_id, action, old_status, new_status, operator, timestamp, notes "
            "FROM payment_logs ORDER BY timestamp DESC LIMIT 10"
        )
        recent = [PaymentLog(*row) for row in cur.fetchall()]

        return {
            "by_status": by_status,
            "paid_today": paid_today,
            "paid_this_month": paid_this_month,
            "overdue_count": len(overdue),
            "overdue_records": overdue,
            "recent_activity": recent,
        }

    def scan_overdue(self, days_threshold: int = 7) -> list[PaymentRecord]:
        threshold_date = (datetime.now() - timedelta(days=days_threshold)).strftime("%Y-%m-%d")
        cur = self._db.execute(
            "SELECT id, payment_no, supplier_name, amount, apply_date, po_number, "
            "status, notes, created_time, updated_time "
            "FROM payment_records WHERE status = '待财务' AND date(updated_time) <= ?",
            (threshold_date,),
        )
        return [PaymentRecord(*row) for row in cur.fetchall()]

    # ── Excel import ──────────────────────────────────────────────────

    def import_from_excel(
        self, filepath: str,
        progress_callback: Callable[[int, int, str], None] | None = None,
    ) -> ImportResult:
        result = ImportResult()
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        if ws is None:
            result.errors.append("无法读取工作表")
            return result

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result.errors.append("工作表为空")
            return result

        header_row, col_map = self._detect_columns(rows)
        if header_row is None:
            result.errors.append("未检测到表头，请确认列名包含: 付款申请单号, 供应商, 金额")
            return result

        data_rows = rows[header_row + 1:]
        if not data_rows:
            result.errors.append("未检测到数据行")
            return result

        total = len(data_rows)
        records_to_insert: list[tuple] = []

        for i, row in enumerate(data_rows):
            if progress_callback:
                progress_callback(i + 1, total, "正在解析...")

            try:
                payment_no = self._cell_str(row, col_map.get("payment_no"))
                supplier = self._cell_str(row, col_map.get("supplier"))
                amount_raw = self._cell_str(row, col_map.get("amount"))
                apply_date = self._cell_str(row, col_map.get("apply_date"))
                po_number = self._cell_str(row, col_map.get("po_number"))
                notes = self._cell_str(row, col_map.get("notes"))

                if not payment_no:
                    result.errors.append(f"第{i + header_row + 2}行: 付款申请单号为空")
                    continue
                if not supplier:
                    result.errors.append(f"第{i + header_row + 2}行: 供应商为空")
                    continue

                try:
                    amount = float(amount_raw.replace(",", "").replace("，", "").replace("¥", "").strip()) if amount_raw else 0
                except ValueError:
                    result.errors.append(f"第{i + header_row + 2}行: 金额「{amount_raw}」无法解析")
                    continue

                existing = self.get_by_payment_no(payment_no)
                if existing:
                    result.skipped += 1
                    continue

                records_to_insert.append((
                    payment_no, supplier, amount,
                    apply_date or datetime.now().strftime("%Y-%m-%d"),
                    po_number, "待审批", notes,
                ))
                result.imported += 1

            except Exception as e:
                result.errors.append(f"第{i + header_row + 2}行: {e}")

        if records_to_insert:
            self._db.executemany(
                "INSERT INTO payment_records (payment_no, supplier_name, amount, apply_date, "
                "po_number, status, notes) VALUES (?, ?, ?, ?, ?, ?, ?)",
                records_to_insert,
            )
            self._db.commit()

        wb.close()
        return result

    @staticmethod
    def _detect_columns(rows: list) -> tuple[int | None, dict[str, int | None]]:
        col_keywords = {
            "payment_no": ["付款申请单号", "申请单号", "付款单号"],
            "supplier": ["供应商"],
            "amount": ["金额", "付款金额"],
            "apply_date": ["申请日期", "日期", "付款日期"],
            "po_number": ["PO号", "PO", "采购订单号", "订单号"],
            "notes": ["备注", "说明"],
        }
        for i in range(min(10, len(rows))):
            row = rows[i]
            if row is None:
                continue
            cells = [str(c).strip() if c is not None else "" for c in row]
            col_map: dict[str, int | None] = {k: None for k in col_keywords}
            for col_idx, cell in enumerate(cells):
                for key, keywords in col_keywords.items():
                    if any(kw in cell for kw in keywords):
                        col_map[key] = col_idx
            if col_map["payment_no"] is not None and col_map["supplier"] is not None:
                return i, col_map
        return None, {}

    @staticmethod
    def _cell_str(row: tuple, col_idx: int | None) -> str:
        if col_idx is None:
            return ""
        if col_idx >= len(row):
            return ""
        val = row[col_idx]
        if val is None:
            return ""
        s = str(val).strip()
        return "" if s in ("None", "") else s

    # ── Excel export ──────────────────────────────────────────────────

    def export_to_excel(self, filepath: str, filters: dict | None = None) -> int:
        records = self.search(filters=filters, limit=10000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "付款记录"
        headers = ["付款申请单号", "供应商", "金额", "申请日期", "PO号", "状态", "备注", "创建时间", "更新时间"]
        ws.append(headers)
        for r in records:
            ws.append([
                r.payment_no, r.supplier_name, r.amount, r.apply_date,
                r.po_number, r.status, r.notes, r.created_time, r.updated_time,
            ])
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        wb.save(filepath)
        return len(records)

    # ── Attachments ───────────────────────────────────────────────────

    def _attachments_dir(self) -> Path:
        d = DB_DIR / "attachments"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def add_attachment(self, payment_id: int, source_path: str) -> PaymentAttachment | None:
        record = self.get_by_id(payment_id)
        if record is None:
            return None

        src = Path(source_path)
        if not src.exists():
            return None

        dest_dir = self._attachments_dir()
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        dest_name = f"{payment_id}_{timestamp}_{src.name}"
        dest_path = dest_dir / dest_name
        shutil.copy2(str(src), str(dest_path))

        cur = self._db.execute(
            "INSERT INTO payment_attachments (payment_id, filename, filepath) VALUES (?, ?, ?)",
            (payment_id, dest_name, str(dest_path)),
        )
        self._db.commit()
        aid = cur.lastrowid
        created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return PaymentAttachment(id=aid, payment_id=payment_id, filename=dest_name, filepath=str(dest_path), created_time=created)

    def get_attachments(self, payment_id: int) -> list[PaymentAttachment]:
        cur = self._db.execute(
            "SELECT id, payment_id, filename, filepath, created_time "
            "FROM payment_attachments WHERE payment_id = ? ORDER BY created_time DESC",
            (payment_id,),
        )
        return [PaymentAttachment(*row) for row in cur.fetchall()]

    def remove_attachment(self, attachment_id: int) -> bool:
        cur = self._db.execute(
            "SELECT filepath FROM payment_attachments WHERE id = ?", (attachment_id,)
        )
        row = cur.fetchone()
        if not row:
            return False
        filepath = row[0]
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except OSError:
            pass
        self._db.execute("DELETE FROM payment_attachments WHERE id = ?", (attachment_id,))
        self._db.commit()
        return True
