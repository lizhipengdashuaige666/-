"""Shared Excel ledger writer — single source of truth for 已发台账.xlsx.

如果台账文件被 Excel 占用，会弹窗提醒用户关闭 Excel 后重试。
"""

import os
import time
import tempfile
import shutil
import ctypes
from datetime import datetime
from pathlib import Path

import openpyxl

LEDGER_PATH = Path(r"D:\采购工作\采购订单\已发\已发台账.xlsx")
HEADERS = ["供应商", "订单号", "文件名称", "类型", "单章合同", "双章合同", "发送日期"]


def _ensure_workbook(excel_path: Path):
    """Return (workbook, worksheet) with headers guaranteed."""
    if excel_path.exists() and excel_path.stat().st_size > 1000:
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("台账")
        for i, h in enumerate(HEADERS, 1):
            existing = ws.cell(row=1, column=i).value
            if existing != h:
                ws.cell(row=1, column=i, value=h)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "台账"
        for i, h in enumerate(HEADERS, 1):
            ws.cell(row=1, column=i, value=h)
    return wb, ws


def _write_atomically(wb, excel_path: Path) -> None:
    """Save to temp file then copy to target (avoids partial writes)."""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        wb.save(tmp.name)
        shutil.copy2(tmp.name, str(excel_path))
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


def append_row(supplier: str, number: str, filename: str,
               business_type: str, date_str: str | None = None,
               excel_path: Path | None = None) -> bool:
    """Append one row. 文件被占用时会弹窗等你关闭 Excel，然后重试。"""
    if excel_path is None:
        excel_path = LEDGER_PATH

    today = date_str or datetime.now().strftime("%Y-%m-%d")

    while True:
        try:
            wb, ws = _ensure_workbook(excel_path)
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=supplier)
            ws.cell(row=row_num, column=2, value=number)
            ws.cell(row=row_num, column=3, value=filename)
            ws.cell(row=row_num, column=4, value=business_type)
            ws.cell(row=row_num, column=5, value="是")
            ws.cell(row=row_num, column=6, value="是")
            ws.cell(row=row_num, column=7, value=today)
            _write_atomically(wb, excel_path)
            return True
        except PermissionError:
            result = ctypes.windll.user32.MessageBoxW(
                0,
                "台账需要记录，请保存并关闭 Excel 中的台账文件，\n然后点「确定」继续。\n点「取消」则跳过本次记录。",
                "台账写入",
                0x00040001 | 0x00000030 | 0x00000100,  # MB_OKCANCEL | MB_ICONWARNING | MB_TOPMOST
            )
            if result != 1:  # 1 = OK, 2 = Cancel
                return False
            time.sleep(0.3)
