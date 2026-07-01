#!/usr/bin/env python3
"""水单自动识别与记账工具 — PySide6 版本。

支持文本型/扫描件 PDF，多笔汇款记录批量提取，结果追加写入水单明细.xlsx。
OCR 使用 PaddleOCR（移动端模型，pdfplumber 优先）。
"""

from __future__ import annotations

import os
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QColor
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QFrame, QHBoxLayout, QLabel,
    QListWidget, QListWidgetItem, QMessageBox, QPushButton,
    QVBoxLayout, QWidget,
)

from unified import style as _style
from unified.ocr import ocr_text_from_pdf

DEFAULT_EXCEL_PATH = r"D:\采购工作\采购订单\已发\水单明细.xlsx"


# ═══════════════════════════════════════════════════════════════════════════
# Extraction helpers
# ═══════════════════════════════════════════════════════════════════════════

MATERIAL_KEYWORDS = [
    "物料", "货款", "采购", "材料", "原料", "零件", "配件",
    "生产", "产品", "货品", "商品", "加工", "模具", "治具",
]


def _match_material(summary_text):
    if not summary_text:
        return False
    for kw in MATERIAL_KEYWORDS:
        if kw in summary_text:
            return True
    return False


def _extract_supplier(text):
    """从文本中提取收款人名称。"""
    m = re.search(r"收款人[：:]?\s*(\S+)", text)
    if m:
        name = re.split(r'[②④⑤⑥⑦⑧⑨CMB]+', m.group(1))[0]
        name = re.sub(r'[a-zA-Z\[\]]+$', '', name).strip()
        if len(name) >= 2:
            return name
    return None


def _extract_payee_account(text):
    """提取收款账号（第一个出现在'收款账号'后面的非付款账号长数字）。"""
    for m in re.finditer(r"收款账号[：:]?\s*(\d{8,25})", text):
        return m.group(1)
    return None


def _extract_summary(segment):
    m = re.search(r"交易摘要[：:]?\s*(\S+)", segment)
    if m:
        val = re.split(r'[②④⑤⑥⑦⑧⑨]', m.group(1))[0]
        val = re.sub(r'[a-zA-Z\[\]]+$', '', val).strip()
        if len(val) >= 2:
            return val
    return None


def _clean_amount(raw, fix_map=None):
    s = re.sub(r"\s+", "", raw).replace(",", "").replace("，", "")
    try:
        return float(s)
    except ValueError:
        pass
    default_fixes = {"O": "0", "S": "5", "Z": "2", "l": "1", "B": "8"}
    for i_val in ["9", "1"]:
        m = dict(default_fixes, I=i_val, i=i_val)
        fixed = s
        for k, v in m.items():
            fixed = fixed.replace(k, v)
        try:
            return float(fixed)
        except ValueError:
            continue
    raise ValueError(f"无法解析金额: {raw}")


# ═══════════════════════════════════════════════════════════════════════════
# Core extraction
# ═══════════════════════════════════════════════════════════════════════════

def _extract_records_from_text(text, pdf_path):
    records = []
    text = re.sub(r"\s+", " ", text)

    # 匹配 CNY 金额行（支持各种括号和分隔符）
    amount_pattern = re.compile(
        r"(?:交易金额[（(]小写[)）]\s*[:：]\s*)?CNY\s*([\d,\s，OIlZSBi]+\.\s?\d{2})",
        re.IGNORECASE,
    )
    amounts = list(amount_pattern.finditer(text))
    if not amounts:
        return records

    # 提取付款人账号（用于排除）
    payer_accounts = set()
    for m in re.finditer(r"付款账号[：:]?\s*(\d{8,})", text):
        payer_accounts.add(m.group(1))

    global_date = datetime.today().strftime("%Y-%m-%d")
    for dm in re.finditer(r"(\d{4}[/-]\d{2}[/-]\d{2})", text):
        try:
            global_date = dm.group(1).replace("/", "-")
            break
        except Exception:
            pass

    amount_positions = [0] + [am.end() for am in amounts]
    for i, am in enumerate(amounts):
        try:
            amount_val = _clean_amount(am.group(1))
        except ValueError:
            continue

        seg_start = amount_positions[i]
        seg_end = am.start()
        segment = text[seg_start:seg_end]
        after_start = am.end()
        after_end = amounts[i + 1].start() if i + 1 < len(amounts) else len(text)
        after_segment = text[after_start:after_end]

        # 在金额前后 1000 字符范围内搜索字段
        context = text[max(0, am.start() - 1000):min(len(text), am.end() + 500)]

        supplier = _extract_supplier(context)
        payee_account = _extract_payee_account(context)

        rec_date = global_date
        for dm in re.finditer(r"(\d{4}[/-]\d{2}[/-]\d{2})", segment):
            try:
                rec_date = dm.group(1).replace("/", "-")
                break
            except Exception:
                pass

        txn_id = None
        m_txn = re.search(r"业务参考号[：:]?\s*(\d{8}\d{10,})", text)
        if not m_txn:
            m_txn = re.search(r"(\d{8}\d{10,})", after_segment)
        if m_txn:
            raw_txn = m_txn.group(1)
            try:
                y_val, mth, d = int(raw_txn[:4]), int(raw_txn[4:6]), int(raw_txn[6:8])
                if 2020 <= y_val <= 2030 and 1 <= mth <= 12 and 1 <= d <= 31:
                    txn_id = raw_txn
            except Exception:
                txn_id = raw_txn

        if not supplier and payee_account:
            supplier = f"(收款账号 {payee_account[-6:]})"

        summary = _extract_summary(segment) or _extract_summary(after_segment) or _extract_summary(context)

        if not _match_material(summary):
            continue

        records.append({
            "amount": amount_val, "supplier": supplier,
            "payee_account": payee_account, "date": rec_date,
            "txn_id": txn_id, "summary": summary,
        })
    return records


def extract_info_from_pdf(pdf_path):
    text_parts = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
    except Exception:
        text_parts = []
    if text_parts:
        full_text = "\n".join(text_parts)
    else:
        try:
            full_text = ocr_text_from_pdf(pdf_path)
        except Exception as e:
            raise RuntimeError(f"PaddleOCR识别失败: {e}")
    if not full_text.strip():
        raise RuntimeError("PDF内容为空，无法提取信息。")
    return _extract_records_from_text(full_text, pdf_path)


# ═══════════════════════════════════════════════════════════════════════════
# Excel output
# ═══════════════════════════════════════════════════════════════════════════

def append_to_excel(records, excel_path=None):
    if excel_path is None:
        excel_path = DEFAULT_EXCEL_PATH
    headers = ["识别日期", "付款日期", "供应商", "收款账号", "付款金额", "交易摘要", "源文件名", "交易流水号"]
    today = datetime.today().strftime("%Y-%m-%d")
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "已付款"
        ws.append(headers)
    else:
        wb = load_workbook(excel_path)
        if "已付款" in wb.sheetnames:
            ws = wb["已付款"]
            existing = [cell.value for cell in ws[1]]
            if existing != headers:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    missing = len(headers) - len(row)
                    for _ in range(missing):
                        row[-1].offset(column=1).value = None
                for i_val, h in enumerate(headers, 1):
                    ws.cell(row=1, column=i_val, value=h)
        else:
            ws = wb.create_sheet("已付款")
            ws.append(headers)
    for rec in records:
        ws.append([
            today, rec.get("date") or "未识别",
            rec.get("supplier") or "未识别", rec.get("payee_account") or "",
            f"{rec['amount']:.2f}" if rec.get("amount") is not None else "未识别",
            rec.get("summary") or "",
            rec.get("filename", ""), rec.get("txn_id") or "",
        ])
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
    wb.save(excel_path)


# ═══════════════════════════════════════════════════════════════════════════
# PySide6 GUI
# ═══════════════════════════════════════════════════════════════════════════

class WatermarkApp(QWidget):
    def __init__(self, parent: QWidget | None = None,
                 shell: object | None = None) -> None:
        super().__init__(parent)
        self._shell = shell
        self.setObjectName("watermarkRoot")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setAcceptDrops(True)
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        header = QLabel("水单自动识别")
        header.setObjectName("toolbarTitle")
        subtitle = QLabel("将银行付款水单 PDF 拖入窗口自动识别汇款记录\n"
                          "仅写入匹配物料关键词的记录 | PaddleOCR 移动端 | pdfplumber 优先")
        subtitle.setObjectName("toolbarSubtitle")
        layout.addWidget(header)
        layout.addWidget(subtitle)

        list_card = QFrame()
        list_card.setObjectName("card")
        list_card.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        if getattr(self._shell, '_theme_overrides', {}).get("_mode") == "light":
            _style.apply_shadow(list_card, elevation=1, is_light=True)
        else:
            _style.apply_shadow(list_card, elevation=1, is_light=False)
        lc = QVBoxLayout(list_card)
        lc.setContentsMargins(14, 12, 14, 12)
        lc.setSpacing(8)

        list_header = QHBoxLayout()
        label = QLabel("待处理 PDF 列表")
        label.setObjectName("cardTitle")
        count_label = QLabel("")
        count_label.setObjectName("sideSubtitle")
        self._count_label = count_label
        list_header.addWidget(label)
        list_header.addStretch()
        list_header.addWidget(count_label)
        lc.addLayout(list_header)

        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.file_list.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)
        remove_action = QAction("移除选中项", self.file_list)
        remove_action.triggered.connect(self._remove_selected)
        self.file_list.addAction(remove_action)
        lc.addWidget(self.file_list, 1)

        layout.addWidget(list_card, 1)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        select_btn = QPushButton("选择文件")
        select_btn.setObjectName("secondaryBtn")
        select_btn.clicked.connect(self._select_files)

        start_btn = QPushButton("开始识别")
        start_btn.setObjectName("primaryBtn")
        start_btn.clicked.connect(self._process_files)

        clear_btn = QPushButton("清空列表")
        clear_btn.setObjectName("dangerButton")
        clear_btn.clicked.connect(self.clear_list)

        btn_row.addWidget(select_btn)
        btn_row.addWidget(start_btn)
        btn_row.addWidget(clear_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self._status = QLabel("就绪")
        self._status.setObjectName("sideSubtitle")
        layout.addWidget(self._status)

    def _select_files(self) -> None:
        from PySide6.QtWidgets import QFileDialog
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择水单PDF", "",
            "PDF文件 (*.pdf);;所有文件 (*.*)")
        self._add_files(paths)

    def add_paths(self, paths: list[Path]) -> None:
        self._add_files([str(p) for p in paths])

    def _add_files(self, paths: list[str]) -> None:
        new = 0
        for p in paths:
            p = os.path.normpath(p)
            if p and p not in {self.file_list.item(i).data(Qt.ItemDataRole.UserRole)
                               for i in range(self.file_list.count())}:
                item = QListWidgetItem(os.path.basename(p))
                item.setData(Qt.ItemDataRole.UserRole, p)
                self.file_list.addItem(item)
                new += 1
        self._refresh_count()

    def clear_list(self) -> None:
        self.file_list.clear()
        self._refresh_count()

    def _remove_selected(self) -> None:
        for item in self.file_list.selectedItems():
            self.file_list.takeItem(self.file_list.row(item))
        self._refresh_count()

    def _refresh_count(self) -> None:
        c = self.file_list.count()
        self._count_label.setText(f"共 {c} 个文件" if c else "")

    def _process_files(self) -> None:
        if self.file_list.count() == 0:
            QMessageBox.information(self, "提示", "请先添加 PDF 文件。")
            return

        self._status.setText("正在识别...")
        self.setCursor(Qt.CursorShape.WaitCursor)
        QApplication.processEvents()

        files = [(self.file_list.item(i).data(Qt.ItemDataRole.UserRole), i)
                 for i in range(self.file_list.count())]
        all_records = []
        errors = []

        for path, _ in files:
            fname = os.path.basename(path)
            try:
                records = extract_info_from_pdf(path)
                for rec in records:
                    rec["filename"] = fname
                all_records.extend(records)
                if not records:
                    errors.append(f"{fname} — 未提取到物料记录")
            except Exception:
                errors.append(f"{fname} — "
                              f"{traceback.format_exc().strip().split(chr(10))[-1]}")

        if all_records:
            try:
                append_to_excel(all_records)
            except Exception:
                errors.append(f"Excel写入失败: "
                              f"{traceback.format_exc().strip().split(chr(10))[-1]}")

        self.setCursor(Qt.CursorShape.ArrowCursor)
        self.clear_list()

        summary = f"成功写入 {len(all_records)} 条物料记录到水单明细"
        if errors:
            detail = "\n".join(errors[-8:])
            summary += f"\n\n{len(errors)} 个问题:\n{detail}"
        self._status.setText(summary.split("\n")[0])
        QMessageBox.information(self, "完成", summary)

    def dragEnterEvent(self, event) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event) -> None:
        paths = [Path(u.toLocalFile()) for u in event.mimeData().urls()]
        pdfs = [str(p) for p in paths if p.suffix.lower() == ".pdf"]
        if pdfs:
            self._add_files(pdfs)
