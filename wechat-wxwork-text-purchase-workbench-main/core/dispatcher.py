from __future__ import annotations

import sys
import time
from pathlib import Path

from adapters import WeChatMessenger, WXWorkMessenger
from core.config import SUPPORTED_PLATFORMS, load_suppliers
from core.logger import logger
from core.parser import parse_document

ADAPTERS = {
    "wechat": WeChatMessenger,
    "wxwork": WXWorkMessenger,
}


def process_file(
    file_path: str | Path,
    supplier_override: str | None = None,
    platform_override: str | None = None,
    chat_name_override: str | None = None,
    skip_open_chat: bool = False,
    open_only: bool = False,
    trust_current_chat: bool = False,
) -> bool:
    path = Path(file_path)
    supplier = (supplier_override or "").strip()
    platform = (platform_override or "").strip()
    chat_name = (chat_name_override or "").strip()
    business_type = "文件"
    number = path.stem

    if not supplier:
        parsed = parse_document(path)
        supplier = parsed.supplier or ""
        business_type = parsed.business_type or business_type
        number = parsed.number or number
        if not supplier:
            print("无法识别供应商，请手动指定", file=sys.stderr)
            logger.error("未识别供应商")
            return False

    if not platform:
        suppliers_map = load_suppliers()
        supplier_config = suppliers_map.get(supplier)
        if not supplier_config:
            print(f"供应商 {supplier} 未配置平台，请在 suppliers.json 中添加", file=sys.stderr)
            logger.error(f"未配置供应商: {supplier}")
            return False
        platform = supplier_config.get("platform", "")
        chat_name = chat_name or supplier_config.get("chat_name", "")

    if platform not in SUPPORTED_PLATFORMS:
        print(f"不支持的平台: {platform}", file=sys.stderr)
        logger.error(f"不支持的平台: {platform}")
        return False

    adapter_cls = ADAPTERS[platform]
    messenger = adapter_cls()
    search_keyword = supplier
    logger.info(
        f"开始处理: 供应商={supplier}, 类型={business_type}, 编号={number}, 平台={platform}, 群聊={search_keyword}"
    )

    if not messenger.activate():
        print(f"无法激活 {messenger.platform} 窗口", file=sys.stderr)
        return False

    if skip_open_chat:
        logger.info("已跳过自动搜索群聊，使用当前打开的聊天窗口发送")
    else:
        for attempt in range(1, 3):
            if messenger.open_chat_by_keyword(search_keyword):
                break
            logger.warning(f"打开聊天失败，第 {attempt}/2 次")
            time.sleep(1)
        else:
            print(f"未找到群聊: {search_keyword}", file=sys.stderr)
            return False

    if open_only:
        if not trust_current_chat:
            print(f"请人工确认已打开正确群聊: {search_keyword}", file=sys.stderr)
            logger.warning(f"未自动确认目标群聊，等待人工确认: {search_keyword}")
            return False
        print(f"已打开群聊: {search_keyword}")
        logger.info(f"已打开群聊，等待用户确认输入框: {search_keyword}")
        return True

    if not messenger.send_file(path):
        print("文件发送失败", file=sys.stderr)
        return False

    import time as _time
    _time.sleep(1)
    _archive_sent_file(path, messenger, supplier)

    print(f"文件已发送至 {supplier} ({messenger.platform})")
    logger.info(f"处理完成: {path}")
    return True


def _archive_sent_file(path: Path, messenger, supplier: str) -> None:
    try:
        sub = "已发/电子档已发"
        archive_dir = path.parent / sub
        archive_dir.mkdir(parents=True, exist_ok=True)
        new_path = archive_dir / path.name
        if new_path.exists():
            from datetime import datetime
            ts = datetime.now().strftime("%H%M%S")
            new_path = archive_dir / f"{path.stem}_{ts}.pdf"
        path.rename(new_path)
        logger.info(f"已归档: {sub}/{new_path.name}")
        _record_to_excel_sent(new_path, supplier)
    except OSError as exc:
        logger.warning(f"归档失败: {path.name} — {exc}")


def _record_to_excel_sent(path: Path, supplier: str) -> None:
    """Record sent file to 台账 Excel."""
    try:
        import openpyxl, tempfile, shutil, os
    except ModuleNotFoundError:
        return
    excel_path = Path(r"D:\采购工作\采购订单\已发\台账测试.xlsx")
    stem = path.stem
    if "合同" not in stem and "合同" not in str(path):
        return
    try:
        try:
            if excel_path.exists() and excel_path.stat().st_size > 1000:
                wb = openpyxl.load_workbook(str(excel_path))
            else:
                raise FileNotFoundError
        except Exception:
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "台账"
            ws.append(["供应商", "订单号", "文件名称", "类型", "单章合同", "双章合同"])

        ws = wb.active
        found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            if row[0].value and str(row[0].value).strip() == supplier:
                found = True
                break

        if not found:
            n = ws.max_row + 1
            ws.cell(row=n, column=1, value=supplier)
            ws.cell(row=n, column=2, value=path.stem)
            ws.cell(row=n, column=3, value=path.name)
            ws.cell(row=n, column=5, value="是")
            ws.cell(row=n, column=6, value="是")

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        wb.save(tmp.name)
        shutil.copy2(tmp.name, str(excel_path))
        os.unlink(tmp.name)
    except PermissionError:
        logger.warning("台账写入失败: 文件被占用")
    except Exception as exc:
        logger.warning(f"台账写入失败: {exc}")
